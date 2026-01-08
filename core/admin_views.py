from __future__ import annotations

import csv
from collections import defaultdict
from datetime import timedelta
from decimal import Decimal, InvalidOperation
import io
import re
from typing import Dict, List

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import (
    Avg,
    Count,
    F,
    FloatField,
    Max,
    Q,
    Sum,
)
from django.db.models import ExpressionWrapper
from django.db.models.functions import TruncMonth, TruncWeek
from django.shortcuts import redirect, render
from django.utils import timezone

from .models import (
    Assessment,
    AssessmentCentre,
    CustomUser,
    ExamSubmission,
    GlobalBusinessRecord,
    Paper,
    PaperBankEntry,
    Qualification,
)

PASS_THRESHOLD = 0.5  # 50% cutoff for pass-rate metrics
COMPLETED_STATUSES = {
    "moderated",
    "etqa_approved",
    "qcto_approved",
    "Released to students",
    "archived",
}

COUNTRY_ALIASES = {
    "usa": "United States",
    "u.s.a": "United States",
    "us": "United States",
    "united states of america": "United States",
    "uk": "United Kingdom",
    "u.k": "United Kingdom",
    "england": "United Kingdom",
    "uae": "United Arab Emirates",
    "drc": "Democratic Republic of the Congo",
    "dr congo": "Democratic Republic of the Congo",
    "sa": "South Africa",
    "s. africa": "South Africa",
}

COUNTRY_CONTINENT_MAP = {
    "south africa": "Africa",
    "nigeria": "Africa",
    "kenya": "Africa",
    "ghana": "Africa",
    "democratic republic of the congo": "Africa",
    "egypt": "Africa",
    "morocco": "Africa",
    "united states": "North America",
    "canada": "North America",
    "mexico": "North America",
    "brazil": "South America",
    "argentina": "South America",
    "chile": "South America",
    "united kingdom": "Europe",
    "ireland": "Europe",
    "germany": "Europe",
    "france": "Europe",
    "netherlands": "Europe",
    "spain": "Europe",
    "italy": "Europe",
    "sweden": "Europe",
    "australia": "Oceania",
    "new zealand": "Oceania",
    "china": "Asia",
    "japan": "Asia",
    "india": "Asia",
    "singapore": "Asia",
    "united arab emirates": "Asia",
    "qatar": "Asia",
}

GLOBAL_DIMENSION_LABELS = {
    "school": "Assessment Centre",
    "country": "Country",
    "continent": "Continent",
}

GLOBAL_DIMENSION_CHOICES = [
    ("school", GLOBAL_DIMENSION_LABELS["school"]),
    ("country", GLOBAL_DIMENSION_LABELS["country"]),
    ("continent", GLOBAL_DIMENSION_LABELS["continent"]),
]

GLOBAL_BUSINESS_FIELD_ALIASES = {
    "school": {"school", "institution", "centre", "center", "campus"},
    "country": {"country", "nation"},
    "continent": {"continent", "region"},
    "learners": {"learners", "students", "enrolled"},
    "submissions": {"submissions", "assessments", "written", "entries"},
    "pass_rate": {"pass_rate", "pass%", "pass %", "success_rate"},
    "average_score": {"average_score", "avg_score", "avg%", "avg %"},
}

GLOBAL_BUSINESS_ALLOWED_EXTENSIONS = (".csv", ".xlsx", ".xlsm", ".xltx", ".xltm")


def _normalize_key(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _coerce_int(value) -> int:
    if value in (None, ""):
        return 0
    try:
        decimal_value = Decimal(str(value))
        return int(decimal_value.to_integral_value())
    except (InvalidOperation, TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0


def _coerce_decimal(value) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        try:
            return Decimal(str(float(value)))
        except (InvalidOperation, TypeError, ValueError):
            return None


def _iter_csv_rows(uploaded_file):
    uploaded_file.seek(0)
    content = uploaded_file.read()
    try:
        text = content.decode("utf-8-sig")
    except AttributeError:
        text = content
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        yield row


def _iter_excel_rows(uploaded_file):
    uploaded_file.seek(0)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Excel uploads require openpyxl to be installed.") from exc

    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    headers = [
        _normalize_key(cell.value) or f"column_{idx}"
        for idx, cell in enumerate(worksheet[1])
    ]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        yield {headers[idx]: value for idx, value in enumerate(row)}


def _parse_global_business_dataset(uploaded_file) -> List[GlobalBusinessRecord]:
    filename = (uploaded_file.name or "").lower()
    if not filename.endswith(GLOBAL_BUSINESS_ALLOWED_EXTENSIONS):
        raise ValueError("Please upload a CSV or Excel file (.csv, .xlsx).")

    if filename.endswith(".csv"):
        raw_rows = list(_iter_csv_rows(uploaded_file))
    else:
        raw_rows = list(_iter_excel_rows(uploaded_file))

    entries: List[GlobalBusinessRecord] = []
    for raw in raw_rows:
        normalized = {
            _normalize_key(key): value for key, value in raw.items() if key is not None
        }

        def pick(field: str):
            for alias in GLOBAL_BUSINESS_FIELD_ALIASES[field]:
                if alias in normalized and normalized[alias] not in (None, ""):
                    return normalized[alias]
            return ""

        school = str(pick("school")).strip()
        if not school:
            continue
        country = str(pick("country")).strip()
        continent = str(pick("continent")).strip()
        learners = _coerce_int(pick("learners"))
        submissions = _coerce_int(pick("submissions"))
        pass_rate = _coerce_decimal(pick("pass_rate"))
        average_score = _coerce_decimal(pick("average_score"))

        entries.append(
            GlobalBusinessRecord(
                school=school,
                country=country,
                continent=continent,
                learners=learners,
                submissions=submissions,
                pass_rate=pass_rate,
                average_score=average_score,
            )
        )

    if not entries:
        raise ValueError("No recognizable rows were found in the uploaded file.")

    return entries


def _resolve_period(period_key: str | None) -> tuple[str, int]:
    period_options = {
        "30d": ("Last 30 days", 30),
        "90d": ("Last 90 days", 90),
        "365d": ("Last 12 months", 365),
    }
    if period_key in period_options:
        label, days = period_options[period_key]
    else:
        label, days = period_options["30d"]
        period_key = "30d"
    return period_key, days


def _safe_int(value: str | None) -> int | None:
    try:
        return int(value) if value else None
    except (TypeError, ValueError):
        return None


def _localize(dt):
    if not dt:
        return None
    if timezone.is_aware(dt):
        return timezone.localtime(dt)
    return dt


def _format_latest_assessment(assessment: Assessment | None) -> str | None:
    if not assessment:
        return None
    parts = [assessment.paper]
    if assessment.eisa_id:
        parts.append(f"({assessment.eisa_id})")
    return " ".join(filter(None, parts))


def _normalize_country_fragment(fragment: str | None) -> str | None:
    if not fragment:
        return None
    cleaned = fragment.strip()
    if not cleaned:
        return None
    lower = cleaned.lower()
    if lower in COUNTRY_ALIASES:
        return COUNTRY_ALIASES[lower]
    return cleaned.title()


def _infer_country_from_location(location: str | None) -> str:
    if not location:
        return "Unknown"
    # Split on common separators and work backwards for the most granular value.
    tokens = [
        part.strip()
        for part in re.split(r"[,/|-]", location)
        if part and part.strip()
    ]
    while tokens:
        candidate = _normalize_country_fragment(tokens[-1])
        if candidate:
            return candidate
        tokens.pop()
    return "Unknown"


def _infer_continent_from_country(country: str) -> str:
    if not country or country == "Unknown":
        return "Unknown"
    return COUNTRY_CONTINENT_MAP.get(country.lower(), "Unknown")


def _aggregate_dimension_rows(rows: List[Dict[str, object]], dimension_key: str) -> List[Dict[str, object]]:
    grouped: dict[str, Dict[str, float | int | str]] = defaultdict(
        lambda: {
            "label": "Unknown",
            "learners": 0,
            "submissions": 0,
            "passed": 0,
            "score_total": 0.0,
            "units": 0,
        }
    )
    for row in rows:
        bucket = row.get(dimension_key) or "Unknown"
        data = grouped[bucket]
        data["label"] = bucket
        data["learners"] = int(data["learners"]) + int(row.get("learners") or 0)
        data["submissions"] = int(data["submissions"]) + int(row.get("submissions") or 0)
        data["passed"] = int(data["passed"]) + int(row.get("passed") or 0)
        data["score_total"] = float(data["score_total"]) + float(row.get("score_total") or 0.0)
        data["units"] = int(data["units"]) + int(row.get("units") or 1)

    aggregated: List[Dict[str, object]] = []
    for bucket, data in grouped.items():
        submissions = int(data["submissions"])
        passed = int(data["passed"])
        score_total = float(data["score_total"])
        aggregated.append(
            {
                "label": data["label"],
                "learners": int(data["learners"]),
                "submissions": submissions,
                "passed": passed,
                "score_total": score_total,
                "avg_score": (score_total / submissions) if submissions else 0.0,
                "pass_rate": (passed / submissions) * 100 if submissions else 0.0,
                "units": int(data["units"]),
            }
        )

    aggregated.sort(key=lambda item: item["learners"], reverse=True)
    return aggregated


def _build_global_business_context(
    *,
    current_start,
    qualification_id: int | None,
    compare_dimension: str,
    compare_values: List[str],
) -> Dict[str, object]:
    safe_dimension = compare_dimension if compare_dimension in GLOBAL_DIMENSION_LABELS else "school"

    uploaded_records = list(GlobalBusinessRecord.objects.all())
    school_rows: List[Dict[str, object]] = []

    if uploaded_records:
        for record in uploaded_records:
            submissions = record.submissions or 0
            pass_rate = float(record.pass_rate or 0.0)
            avg_score = float(record.average_score or 0.0)
            passed = int(round(submissions * (pass_rate / 100.0)))
            score_total = submissions * avg_score
            country = record.country.strip() if record.country else "Unknown"
            continent = record.continent.strip() if record.continent else "Unknown"
            school_rows.append(
                {
                    "label": record.school,
                    "school": record.school,
                    "country": country or "Unknown",
                    "continent": continent or "Unknown",
                    "learners": record.learners or 0,
                    "submissions": submissions,
                    "passed": passed,
                    "score_total": score_total,
                    "avg_score": avg_score,
                    "pass_rate": pass_rate,
                    "units": 1,
                }
            )
    else:
        learner_qs = CustomUser.objects.filter(
            role="learner",
            assessment_centre__isnull=False,
            created_at__gte=current_start,
        )
        if qualification_id:
            learner_qs = learner_qs.filter(qualification_id=qualification_id)

        learner_rows = learner_qs.values("assessment_centre_id").annotate(total=Count("id"))
        learner_map = {
            row["assessment_centre_id"]: row["total"]
            for row in learner_rows
            if row["assessment_centre_id"]
        }

        score_expression = ExpressionWrapper(
            F("marks") * 100.0 / F("total_marks"),
            output_field=FloatField(),
        )
        submission_qs = ExamSubmission.objects.filter(
            student__assessment_centre__isnull=False,
            submitted_at__gte=current_start,
            total_marks__gt=0,
            marks__isnull=False,
        )
        if qualification_id:
            submission_qs = submission_qs.filter(assessment__qualification_id=qualification_id)

        submission_rows = (
            submission_qs.values("student__assessment_centre_id")
            .annotate(
                submitted=Count("id"),
                passed=Count(
                    "id",
                    filter=Q(marks__gte=F("total_marks") * PASS_THRESHOLD),
                ),
                score_total=Sum(score_expression),
            )
        )
        submission_map = {
            row["student__assessment_centre_id"]: {
                "submitted": row["submitted"],
                "passed": row["passed"],
                "score_total": float(row["score_total"] or 0.0),
            }
            for row in submission_rows
            if row["student__assessment_centre_id"]
        }

        centre_ids = set(learner_map.keys()) | set(submission_map.keys())
        if not centre_ids:
            return {
                "global_summary": {
                    "schools": 0,
                    "countries": 0,
                    "continents": 0,
                    "total_learners": 0,
                    "active_submissions": 0,
                    "top_school": None,
                    "top_school_rate": None,
                },
                "global_dimension_options": {key: [] for key in GLOBAL_DIMENSION_LABELS},
                "global_compare_dimension": safe_dimension,
                "global_compare_dimension_label": GLOBAL_DIMENSION_LABELS[safe_dimension],
                "global_compare_values": [],
                "global_compare_rows": [],
                "global_compare_chart_data": [],
                "global_dimension_choices": GLOBAL_DIMENSION_CHOICES,
                "global_active_dimension_options": [],
                "global_has_data": False,
            }

        centres = AssessmentCentre.objects.filter(id__in=centre_ids).values("id", "name", "location")
        centre_map = {row["id"]: row for row in centres}

        for centre_id, centre in centre_map.items():
            learners = learner_map.get(centre_id, 0)
            submission_stats = submission_map.get(centre_id, {})
            submissions = submission_stats.get("submitted") or 0
            passed = submission_stats.get("passed") or 0
            score_total = submission_stats.get("score_total") or 0.0
            avg_score = (score_total / submissions) if submissions else 0.0
            pass_rate = (passed / submissions) * 100 if submissions else 0.0
            country = _infer_country_from_location(centre.get("location"))
            continent = _infer_continent_from_country(country)

            school_rows.append(
                {
                    "label": centre["name"],
                    "school": centre["name"],
                    "country": country,
                    "continent": continent,
                    "learners": learners,
                    "submissions": submissions,
                    "passed": passed,
                    "score_total": score_total,
                    "avg_score": avg_score,
                    "pass_rate": pass_rate,
                    "units": 1,
                }
            )

    school_rows.sort(key=lambda row: row["learners"], reverse=True)
    dimension_data = {
        "school": school_rows,
        "country": _aggregate_dimension_rows(school_rows, "country"),
        "continent": _aggregate_dimension_rows(school_rows, "continent"),
    }
    dimension_options = {}
    for key, rows in dimension_data.items():
        labels = [row["label"] for row in rows if row["label"]]
        dimension_options[key] = sorted(set(labels), key=lambda label: label.lower())

    dimension_rows = dimension_data.get(safe_dimension, [])
    available_labels = [row["label"] for row in dimension_rows]
    selected_values = [value for value in compare_values if value in available_labels]
    if not selected_values and available_labels:
        selected_values = available_labels[:3]

    compare_rows = [row for row in dimension_rows if row["label"] in selected_values]
    chart_data = [
        {
            "label": row["label"],
            "learners": int(row["learners"]),
            "submissions": int(row["submissions"]),
            "pass_rate": round(float(row["pass_rate"] or 0.0), 1),
            "avg_score": round(float(row["avg_score"] or 0.0), 1),
        }
        for row in compare_rows
    ]

    countries = {row["country"] for row in school_rows if row["country"] and row["country"] != "Unknown"}
    continents = {row["continent"] for row in school_rows if row["continent"] and row["continent"] != "Unknown"}
    top_school = max(
        (row for row in school_rows if row["submissions"]),
        key=lambda item: item["pass_rate"],
        default=None,
    )
    summary = {
        "schools": len(school_rows),
        "countries": len(countries),
        "continents": len(continents),
        "total_learners": sum(int(row["learners"]) for row in school_rows),
        "active_submissions": sum(int(row["submissions"]) for row in school_rows),
        "top_school": top_school["label"] if top_school else None,
        "top_school_rate": round(float(top_school["pass_rate"]), 1) if top_school else None,
    }

    return {
        "global_summary": summary,
        "global_dimension_options": dimension_options,
        "global_compare_dimension": safe_dimension,
        "global_compare_dimension_label": GLOBAL_DIMENSION_LABELS[safe_dimension],
        "global_compare_values": selected_values,
        "global_compare_rows": compare_rows,
        "global_compare_chart_data": chart_data,
        "global_dimension_choices": GLOBAL_DIMENSION_CHOICES,
        "global_active_dimension_options": dimension_options.get(safe_dimension, []),
        "global_has_data": bool(school_rows),
    }


@login_required
def administrator_analytics_dashboard(request):
    default_tab = getattr(request, "default_dashboard_tab", "analytics")
    active_tab = request.GET.get("tab") or default_tab
    period_key, period_days = _resolve_period(request.GET.get("period"))
    qualification_id = _safe_int(request.GET.get("qualification"))
    paper_type = request.GET.get("paper_type") or ""

    now = timezone.now()
    current_start = now - timedelta(days=period_days)
    previous_start = current_start - timedelta(days=period_days)

    assessments_base = Assessment.objects.select_related("qualification", "paper_link")
    exam_submissions_base = ExamSubmission.objects.select_related("assessment__qualification")
    learner_users = CustomUser.objects.filter(role="learner")

    if qualification_id:
        assessments_base = assessments_base.filter(qualification_id=qualification_id)
        exam_submissions_base = exam_submissions_base.filter(assessment__qualification_id=qualification_id)
        learner_users = learner_users.filter(qualification_id=qualification_id)

    if paper_type:
        assessments_base = assessments_base.filter(paper_type=paper_type)
        exam_submissions_base = exam_submissions_base.filter(assessment__paper_type=paper_type)

    assessments_current = assessments_base.filter(created_at__gte=current_start)
    assessments_previous = assessments_base.filter(created_at__gte=previous_start, created_at__lt=current_start)

    submissions_current = exam_submissions_base.filter(submitted_at__gte=current_start)
    submissions_previous = exam_submissions_base.filter(submitted_at__gte=previous_start, submitted_at__lt=current_start)

    total_learners = learner_users.count()
    active_qualifications = (
        learner_users.exclude(qualification__isnull=True).values("qualification").distinct().count()
    )

    current_completed = assessments_current.filter(status__in=COMPLETED_STATUSES).count()
    previous_completed = assessments_previous.filter(status__in=COMPLETED_STATUSES).count()
    completed_delta = None
    if previous_completed:
        completed_delta = ((current_completed - previous_completed) / previous_completed) * 100

    pass_summary_current = submissions_current.filter(
        marks__isnull=False, total_marks__gt=0
    ).aggregate(
        total=Count("id"),
        passed=Count("id", filter=Q(marks__gte=F("total_marks") * PASS_THRESHOLD)),
    )
    pass_summary_previous = submissions_previous.filter(
        marks__isnull=False, total_marks__gt=0
    ).aggregate(
        total=Count("id"),
        passed=Count("id", filter=Q(marks__gte=F("total_marks") * PASS_THRESHOLD)),
    )

    current_total = pass_summary_current["total"] or 0
    current_passed = pass_summary_current["passed"] or 0
    previous_total = pass_summary_previous["total"] or 0
    previous_passed = pass_summary_previous["passed"] or 0

    overall_pass_rate = (current_passed / current_total) * 100 if current_total else 0.0
    pass_rate_delta = None
    if previous_total:
        previous_rate = (previous_passed / previous_total) * 100
        pass_rate_delta = overall_pass_rate - previous_rate

    paper_ids = assessments_current.filter(
        paper_link__isnull=False
    ).values_list("paper_link_id", flat=True).distinct()
    papers_qs = Paper.objects.filter(id__in=paper_ids)
    randomized_papers = papers_qs.filter(is_randomized=True).count()
    total_papers = papers_qs.count()
    randomized_share = (randomized_papers / total_papers) * 100 if total_papers else 0.0

    metrics = {
        "total_learners": total_learners,
        "active_qualifications": active_qualifications,
        "completed_assessments": current_completed,
        "completed_assessments_delta": completed_delta,
        "overall_pass_rate": overall_pass_rate,
        "pass_rate_delta": pass_rate_delta,
        "randomized_papers": randomized_papers,
        "randomized_share": randomized_share,
    }

    pass_rate_rows = (
        submissions_current.filter(marks__isnull=False, total_marks__gt=0)
        .values("assessment__qualification__name")
        .annotate(
            total=Count("id"),
            passed=Count("id", filter=Q(marks__gte=F("total_marks") * PASS_THRESHOLD)),
        )
        .order_by("-total")
    )
    pass_rate_data: List[Dict[str, float | str]] = []
    for row in pass_rate_rows:
        total = row["total"] or 0
        passed = row["passed"] or 0
        if not total:
            continue
        pass_rate_data.append(
            {
                "qualification": row["assessment__qualification__name"] or "Unassigned",
                "pass_rate": round((passed / total) * 100, 1),
            }
        )

    trunc_fn = TruncWeek if period_days <= 90 else TruncMonth
    completion_trend_rows = (
        assessments_base.filter(status__in=COMPLETED_STATUSES, created_at__gte=current_start)
        .annotate(period=trunc_fn("created_at"))
        .values("period")
        .annotate(completed=Count("id"))
        .order_by("period")
    )
    completion_trend_data: List[Dict[str, str | int]] = []
    for row in completion_trend_rows:
        period = row["period"]
        if not period:
            continue
        localized = _localize(period)
        label = localized.strftime("%b %d, %Y") if period_days <= 90 else localized.strftime("%b %Y")
        completion_trend_data.append({"label": label, "completed": row["completed"]})

    assessment_type_rows = (
        assessments_current.values("paper_type").annotate(total=Count("id")).order_by("paper_type")
    )
    assessment_type_breakdown: Dict[str, int] = {}
    type_labels = dict(Assessment.PAPER_TYPE_CHOICES)
    for row in assessment_type_rows:
        key = row["paper_type"] or "unknown"
        label = type_labels.get(key, key.title())
        assessment_type_breakdown[label] = row["total"]

    enrollment_rows = (
        learner_users.values("qualification__name")
        .annotate(total=Count("id"))
        .order_by("-total", "qualification__name")
    )
    enrollment_by_course_data = [
        {
            "qualification": row["qualification__name"] or "Unassigned",
            "learners": row["total"],
        }
        for row in enrollment_rows
    ][:12]

    learners_map = {
        row["qualification_id"]: row["total"]
        for row in learner_users.values("qualification_id").annotate(total=Count("id"))
    }

    completed_learners_map = {
        row["assessment__qualification_id"]: row["unique_learners"]
        for row in submissions_current.filter(student_number__isnull=False)
        .values("assessment__qualification_id")
        .annotate(unique_learners=Count("student_number", distinct=True))
    }

    score_expression = ExpressionWrapper(
        F("marks") * 100.0 / F("total_marks"),
        output_field=FloatField(),
    )
    avg_score_rows = (
        submissions_current.filter(total_marks__gt=0, marks__isnull=False)
        .annotate(score_pct=score_expression)
        .values("assessment__qualification_id")
        .annotate(avg_score=Avg("score_pct"))
    )
    avg_score_map = {
        row["assessment__qualification_id"]: row["avg_score"] for row in avg_score_rows
    }

    pass_rate_map = {
        row["assessment__qualification_id"]: (
            (row["passed"] / row["total"]) * 100 if row["total"] else None
        )
        for row in (
            submissions_current.filter(total_marks__gt=0, marks__isnull=False)
            .values("assessment__qualification_id")
            .annotate(
                total=Count("id"),
                passed=Count("id", filter=Q(marks__gte=F("total_marks") * PASS_THRESHOLD)),
            )
        )
    }

    assessment_counts = (
        assessments_current.values("qualification_id", "qualification__name")
        .annotate(
            written_count=Count("id", filter=Q(paper_type="admin_upload")),
            randomized_count=Count("id", filter=Q(paper_type="randomized")),
            total=Count("id"),
            latest_created=Max("created_at"),
        )
        .order_by("qualification__name")
    )

    latest_assessment_map: Dict[int | None, str | None] = {}
    for assessment in assessments_current.order_by("-created_at"):
        qid = assessment.qualification_id
        if qid not in latest_assessment_map:
            latest_assessment_map[qid] = _format_latest_assessment(assessment)

    course_statistics: List[Dict[str, object]] = []
    for row in assessment_counts:
        qid = row["qualification_id"]
        course_statistics.append(
            {
                "qualification": row["qualification__name"] or "Unassigned",
                "total_learners": learners_map.get(qid, 0),
                "completed_learners": completed_learners_map.get(qid, 0),
                "pass_rate": pass_rate_map.get(qid),
                "average_score": avg_score_map.get(qid),
                "written_count": row["written_count"],
                "randomized_count": row["randomized_count"],
                "latest_assessment": latest_assessment_map.get(qid),
            }
        )

    compare_dimension = request.GET.get("compare_dim") or "school"
    compare_values = request.GET.getlist("compare_values")
    global_context = _build_global_business_context(
        current_start=current_start,
        qualification_id=qualification_id,
        compare_dimension=compare_dimension,
        compare_values=compare_values,
    )

    context = {
        "metrics": metrics,
        "qualifications": Qualification.objects.all().order_by("name"),
        "assessment_type_choices": Assessment.PAPER_TYPE_CHOICES,
        "period_options": [
            ("30d", "Last 30 days"),
            ("90d", "Last 90 days"),
            ("365d", "Last 12 months"),
        ],
        "pass_rate_data": pass_rate_data,
        "completion_trend_data": completion_trend_data,
        "assessment_type_breakdown": assessment_type_breakdown,
        "enrollment_by_course_data": enrollment_by_course_data,
        "course_statistics": course_statistics,
        "active_page": "analytics-dashboard",
    }
    context.update(global_context)
    context["active_tab"] = active_tab
    context["request"] = request
    return render(request, "core/administrator/dashboards.html", context)


@login_required
def administrator_global_business_dashboard(request):
    setattr(request, "default_dashboard_tab", "global")
    return administrator_analytics_dashboard(request)


@login_required
def global_business_upload_dashboard(request):
    if not request.user.is_staff:
        raise PermissionDenied("Administrator access is required for this page.")

    records_qs = GlobalBusinessRecord.objects.order_by("school")
    record_count = records_qs.count()
    last_uploaded_at = records_qs.aggregate(last=Max("uploaded_at"))["last"]

    if request.method == "POST":
        dataset_file = request.FILES.get("dataset_file")
        if not dataset_file:
            messages.error(request, "Select a CSV or Excel file before uploading.")
            return redirect("global_business_upload")

        try:
            entries = _parse_global_business_dataset(dataset_file)
        except ValueError as exc:
            messages.error(request, str(exc))
        else:
            with transaction.atomic():
                GlobalBusinessRecord.objects.all().delete()
                GlobalBusinessRecord.objects.bulk_create(entries)
            messages.success(
                request,
                f"Uploaded {len(entries)} row(s) to the Global Business dashboard.",
            )
            return redirect("global_business_upload")

    context = {
        "records": records_qs,
        "record_count": record_count,
        "last_uploaded_at": last_uploaded_at,
        "expected_headers": [
            "school",
            "country",
            "continent",
            "learners",
            "submissions",
            "pass_rate",
            "average_score",
        ],
        "active_page": "global-business-upload",
    }
    context["request"] = request
    return render(request, "core/administrator/global_business_dashboard.html", context)


@login_required
def administrator_paperbank(request):
    qualification_id = _safe_int(request.GET.get("qualification"))
    paper_type = request.GET.get("paper_type") or ""
    status_filter = request.GET.get("status") or ""
    query = (request.GET.get("q") or "").strip()

    entries_qs = PaperBankEntry.objects.select_related(
        "assessment",
        "assessment__qualification",
        "assessment__paper_link",
        "assessment__paper_link__created_by",
    ).order_by("-created_at")

    if qualification_id:
        entries_qs = entries_qs.filter(assessment__qualification_id=qualification_id)

    if paper_type:
        entries_qs = entries_qs.filter(assessment__paper_type=paper_type)

    if status_filter:
        entries_qs = entries_qs.filter(assessment__status=status_filter)

    if query:
        entries_qs = entries_qs.filter(
            Q(assessment__eisa_id__icontains=query)
            | Q(assessment__paper__icontains=query)
        )

    total_entries = entries_qs.count()
    aggregates = entries_qs.aggregate(
        randomized_entries=Count(
            "id",
            filter=Q(assessment__paper_link__is_randomized=True)
            | Q(assessment__paper_type="randomized"),
        ),
        memos_available=Count(
            "id",
            filter=Q(assessment__memo__isnull=False)
            | Q(assessment__memo_file__isnull=False),
        ),
        latest_upload_at=Max("created_at"),
        average_total_marks=Avg(
            "assessment__paper_link__total_marks",
            filter=Q(assessment__paper_link__total_marks__isnull=False),
            output_field=FloatField(),
        ),
    )

    randomized_entries = aggregates["randomized_entries"] or 0
    memos_available = aggregates["memos_available"] or 0
    latest_upload_at = aggregates["latest_upload_at"]
    average_total_marks = aggregates["average_total_marks"]

    paperbank_stats = {
        "total_entries": total_entries,
        "randomized_entries": randomized_entries,
        "randomized_share": (randomized_entries / total_entries) * 100 if total_entries else 0.0,
        "memos_available": memos_available,
        "memos_share": (memos_available / total_entries) * 100 if total_entries else 0.0,
        "latest_upload_at": _localize(latest_upload_at).strftime("%Y-%m-%d %H:%M") if latest_upload_at else None,
        "average_total_marks": average_total_marks,
    }

    uploads_by_month_rows = (
        entries_qs.annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )
    uploads_by_month = [
        {"label": _localize(row["month"]).strftime("%b %Y"), "count": row["count"]}
        for row in uploads_by_month_rows
        if row["month"]
    ]

    distribution_rows = (
        entries_qs.values("assessment__qualification__name")
        .annotate(count=Count("id"))
        .order_by("-count", "assessment__qualification__name")
    )
    paperbank_distribution = [
        {
            "qualification": row["assessment__qualification__name"] or "Unassigned",
            "count": row["count"],
        }
        for row in distribution_rows
    ]

    context = {
        "qualifications": Qualification.objects.all().order_by("name"),
        "paper_type_choices": Assessment.PAPER_TYPE_CHOICES,
        "status_options": Assessment.STATUS_CHOICES,
        "paperbank_stats": paperbank_stats,
        "uploads_by_month": uploads_by_month,
        "paperbank_distribution": paperbank_distribution,
        "paper_bank_entries": entries_qs,
        "active_page": "paper-bank",
    }
    context["request"] = request
    return render(request, "core/administrator/paperbank.html", context)
